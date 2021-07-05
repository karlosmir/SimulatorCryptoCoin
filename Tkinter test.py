# Autor: Carlos Mir Martínez

# Librerias
# Json para almacenar e intercambiar datos
import json
import time
#widget
import tkinter
from tkinter import *
from tkinter import ttk
import tkinter as tk
#escribir excell
import openpyxl
from datetime import datetime

# bs4 es la libreria para sacar los datos de htmls y xml
from bs4 import BeautifulSoup

# Requests se utiliza para realizar peticioes a la pagina que vamos a extraer los datos
import requests

# Un agente de usuario es una aplicación informática que funciona
# como cliente en un protocolo de red; el nombre se aplica generalmente
# para referirse a aquellas aplicaciones que acceden a la World Wide Web.
# Los agentes de usuario que se conectan a la Web pueden ser desde
# navegadores web hasta los web crawler de los buscadores
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}

#Variables globales
#Diccionario Global c[simbolo] = valor
c = {}
monedero = 0
d = ['BTC', 'ETH', 'USDT', 'BNB', 'ADA']
l = ['pid-1057391-last', 'pid-1061443-last', 'pid-1061453-last','pid-1061448-last', 'pid-1062537-last']

#Funcion que apunta a la url de investing
def Investing():
    global l
    global d
    # Nuestra url de criptomonedas que vamos a utilizar para sacar informacion
    url='https://es.investing.com/crypto/currencies'

    # La peticion que realizamos a la url
    content=requests.get(url,headers=headers)
    soup=BeautifulSoup(content.text,'html.parser')
    j = 0
    for i in l:
        valor = soup.find_all("a", attrs={"class":i})[0].get_text()
        valor = valor.replace(".", "")
        valor = valor.replace(",", ".")
        valor = "{:.2f}".format(float(valor) / EURO)
        c[d[j]]=valor
        j = j + 1

def mostrar():
    Investing()
    lbl.configure(text=c["BTC"])
    fecha.configure(text=time.strftime("%H:%M:%S"))

def mostrar1():
    Investing()
    lbl1.configure(text=c["ETH"])
    fecha1.configure(text=time.strftime("%H:%M:%S"))

def mostrar2():
    Investing()
    lbl2.configure(text=c["USDT"])
    fecha2.configure(text=time.strftime("%H:%M:%S"))

def mostrar3():
    Investing()
    lbl3.configure(text=c["BNB"])
    fecha3.configure(text=time.strftime("%H:%M:%S"))

def mostrar4():
    Investing()
    lbl4.configure(text=c["ADA"])
    fecha4.configure(text=time.strftime("%H:%M:%S"))

def mostrar_todas():
    Investing()
    lbl.configure(text=c["BTC"])
    fecha.configure(text=time.strftime("%H:%M:%S"))
    lbl1.configure(text=c["ETH"])
    fecha1.configure(text=time.strftime("%H:%M:%S"))
    lbl2.configure(text=c["USDT"])
    fecha2.configure(text=time.strftime("%H:%M:%S"))
    lbl3.configure(text=c["BNB"])
    fecha3.configure(text=time.strftime("%H:%M:%S"))
    lbl4.configure(text=c["ADA"])
    fecha4.configure(text=time.strftime("%H:%M:%S"))

def display_selected(choice):
    choice = variable.get()
    Simbolo_C.configure(text=choice)

def inversion():
    global procesos
    global fila
    global id
    global monedero

    wb_m = openpyxl.load_workbook(('Tkinter.xlsx'))
    hoja_m = wb_m.active
    cell = hoja_m['A%d' % fila]
    cond = True

    Investing()
    simbolo = variable.get()
    valor_cripto = c[simbolo]
    valor = float(valor_cripto)
    valor_monedero = float(Input.get())
    monedero = monedero - valor_monedero
    BalanceOut.configure(text=monedero)
    hoja_m['P2'] = monedero

    cantidad_cripto = valor_monedero / valor
    cantidad_cripto = "{:.14f}".format(cantidad_cripto)
    listBox.insert(tkinter.END,  simbolo ,str(cantidad_cripto) , str(valor_monedero) + " - inversion €" ,str(valor) + " - valor moneda actual €",time.strftime("%H:%M:%S"))
    listBox.insert(tkinter.END, " " )

    while (cond):
        if cell.value is None:
            hoja_m['A%d' % (fila)] = id
            hoja_m['B%d' % (fila)] = simbolo
            hoja_m['C%d' % (fila)] = valor_cripto
            hoja_m['D%d' % (fila)] = valor_monedero
            hoja_m['E%d' % (fila)] = cantidad_cripto
            hoja_m['F%d' % (fila)] = time.strftime("%b %d %Y %H:%M:%S")
            id = id + 1
            fila + 1
            cond = False

        else:
            id = hoja_m['A%d' % (fila)].value + 1
            fila = fila + 1
            cell = hoja_m['A%d' % fila]

    wb_m.save('Tkinter.xlsx')
    procesos = procesos + [str(cantidad_cripto) + " " + simbolo]
    lista_desplegable['values'] = (procesos)
    Output.configure(text=cantidad_cripto)

def vender():
    global procesos2
    global fila2
    global id2
    global monedero

    wb_m = openpyxl.load_workbook(('Tkinter.xlsx'))
    hoja_m = wb_m.active
    cell = hoja_m['I%d' % fila2]
    cond2 = True

    Investing()
    cantidad_moneda= float(Inputv.get())
    simbolo = variable.get()
    valor_cripto = c[simbolo]
    valor = float(valor_cripto)
    total =valor * cantidad_moneda
    monedero = float(monedero) + float(total)
    total = "{:.6f}".format(total)

    BalanceOut.configure(text=monedero)
    hoja_m['P2'] = monedero

    listBox2.insert(tkinter.END,  simbolo ,str(cantidad_moneda) , str(total) + " - resultado venta €" ,str(valor) + " - valor moneda actual €",time.strftime("%H:%M:%S"))
    listBox2.insert(tkinter.END, " ")

    while (cond2):
        if cell.value is None:
            hoja_m['I%d' % (fila2)] = id2
            hoja_m['J%d' % (fila2)] = simbolo
            hoja_m['K%d' % (fila2)] = valor
            hoja_m['L%d' % (fila2)] = total
            hoja_m['M%d' % (fila2)] = cantidad_moneda
            hoja_m['N%d' % (fila2)] = time.strftime("%b %d %Y %H:%M:%S")
            id2 = id2 + 1
            fila2 + 1
            cond2 = False

        else:
            id2 = hoja_m['I%d' % (fila2)].value + 1
            fila2 = fila2 + 1
            cell = hoja_m['I%d' % fila2]

    wb_m.save('Tkinter.xlsx')

    procesos2 = procesos2 + [str(cantidad_moneda) + " " + simbolo]
    lista_desplegable2['values'] = (procesos2)
    Output2.configure(text=total)

def Limpiar():
    lista_desplegable['values'] = ()
    lista_desplegable2['values'] = ()

def Añadir():
    global monedero
    monedero = monedero + float(InputM.get())
    InputM.configure(state='disabled')
    AñadirSaldo.configure(state='disabled')
    BalanceOut.configure(text=monedero)

# MAIN
window = Tk()
EURO = 1.19
wb = openpyxl.load_workbook(('Tkinter.xlsx'))
from openpyxl.styles import Font
hoja = wb.active

hoja['A1'].font = Font(size=10, bold=True)
hoja['B1'].font = Font(size=10, bold=True)
hoja['C1'].font = Font(size=10, bold=True)
hoja['D1'].font = Font(size=10, bold=True)
hoja['E1'].font = Font(size=10, bold=True)
hoja['F1'].font = Font(size=10, bold=True)

hoja['I1'].font = Font(size=10, bold=True)
hoja['J1'].font = Font(size=10, bold=True)
hoja['K1'].font = Font(size=10, bold=True)
hoja['L1'].font = Font(size=10, bold=True)
hoja['M1'].font = Font(size=10, bold=True)
hoja['N1'].font = Font(size=10, bold=True)

hoja['P1'].font = Font(size=10, bold=True)

hoja['A1'] = "ID COMPRA"
hoja['B1'] = "SIMBOLO"
hoja['C1'] = "VALOR COIN"
hoja['D1'] = "INVERSION"
hoja['E1'] = "CANTIDAD COIN"
hoja['F1'] = "FECHA"

hoja['I1'] = "ID VENTA"
hoja['J1'] = "SIMBOLO"
hoja['K1'] = "VALOR COIN"
hoja['L1'] = "RESULTADO VENTA"
hoja['M1'] = "CANTIDAD COIN"
hoja['N1'] = "FECHA"

hoja['P1'] = "MONEDERO"
celda = hoja['P2']
if celda.value is None:
    monedero = 0
else:
    monedero = hoja['P2'].value

fila = 2
fila2 = 2
id2 = 1
id = 1

wb.save('Tkinter.xlsx')

#Ajustes
window.geometry('700x400')
window.title("Criptomonedas")
tab_control = ttk.Notebook(window )
tab1 = ttk.Frame(tab_control)
tab2 = ttk.Frame(tab_control)
tab3 = ttk.Frame(tab_control)
tab4 = ttk.Frame(tab_control)
tab_control.add(tab1, text='Valores' )
tab_control.add(tab2, text='Transacciones')
tab_control.add(tab3, text='Historial de Compras')
tab_control.add(tab4, text='Historial de Ventas')

#Pestaña 2
procesos = ['Compras']
lista_desplegable = ttk.Combobox(tab2,width=25, state="readonly")
lista_desplegable.place(x=330,y=0)
lista_desplegable['values']= ()

procesos2 = ['Ventas']
lista_desplegable2 = ttk.Combobox(tab2,width=25,state="readonly")
lista_desplegable2.place(x=330,y=25)
lista_desplegable2['values']= ()

#Pestaña 3
listBox = Listbox(tab3, width=50, height=30)
listBox.pack(side="left", fill="both",expand=False)
listProcesos = []
i=1
for i in [10]:
    listProcesos += [[i, '', '', '', '', '']]

#Pestaña 4
listBox2 = Listbox(tab4, width=50, height=30)
listBox2.pack(side="left", fill="both",expand=False)
listProcesos2 = []
j=1
for j in [10]:
    listProcesos2 += [[i, '', '', '', '', '']]

# Windows general
simbolos = ['BTC', 'ETH', 'USDT', 'BNB', 'ADA']
variable=StringVar()
variable.set(simbolos[4])
dropdown=OptionMenu(
    window,
    variable,
    *simbolos,
    command=display_selected
)
dropdown.pack(expand=False)

# Pestaña 1
btn = Button(tab1, text="BTC", bg="green", fg="black", command=mostrar, width=10 , font='Helvetica 9 bold')
btn.grid(column=0, row=0)
lbl = Label(tab1, bg="green", fg="black",width=10 , font='Helvetica 9 bold')
lbl.grid(column=1, row=0)
fecha = Label(tab1, bg="pink", fg="black",width=10, font='Helvetica 9 bold')
fecha.grid(column=2, row=0)

btn1 = Button(tab1, text="ETH", bg="green", fg="black", command=mostrar1,width=10, font='Helvetica 9 bold')
btn1.grid(column=0, row=1)
lbl1 = Label(tab1, bg="green", fg="black",width=10, font='Helvetica 9 bold')
lbl1.grid(column=1, row=1)
fecha1 = Label(tab1, bg="pink", fg="black",width=10, font='Helvetica 9 bold')
fecha1.grid(column=2, row=1)

btn2 = Button(tab1, text="USDT", bg="green", fg="black", command=mostrar2,width=10, font='Helvetica 9 bold')
btn2.grid(column=0, row=2)
lbl2 = Label(tab1, bg="green", fg="black",width=10, font='Helvetica 9 bold')
lbl2.grid(column=1, row=2)
fecha2 = Label(tab1, bg="pink", fg="black",width=10, font='Helvetica 9 bold')
fecha2.grid(column=2, row=2)

btn3 = Button(tab1, text="BNB", bg="green", fg="black", command=mostrar3,width=10, font='Helvetica 9 bold')
btn3.grid(column=0, row=3)
lbl3 = Label(tab1, bg="green", fg="black",width=10, font='Helvetica 9 bold')
lbl3.grid(column=1, row=3)
fecha3 = Label(tab1, bg="pink", fg="black",width=10, font='Helvetica 9 bold')
fecha3.grid(column=2, row=3)

btn4 = Button(tab1, text="ADA", bg="green", fg="black", command=mostrar4,width=10, font='Helvetica 9 bold')
btn4.grid(column=0, row=4)
lbl4 = Label(tab1, bg="green", fg="black",width=10, font='Helvetica 9 bold')
lbl4.grid(column=1, row=4)
fecha4 = Label(tab1, bg="pink", fg="black",width=10, font='Helvetica 9 bold')
fecha4.grid(column=2, row=4)

Mostrar_T = Button(tab1, text ="Mostrar Todas", bg="green", fg="black", command=mostrar_todas, width=15, fon='Helvetica 9 bold')
Mostrar_T.grid(column=4,row=0)

# Pestaña 2
Monedero = Label(tab2, text="Inversion" , bg="pink" , fg="black",width= 18, font='Helvetica 9 bold')
Monedero.grid(column=0, row=0)
Resultado = Label(tab2, text="Resultado Cripto", bg="pink", fg="black",width=18, font='Helvetica 9 bold')
Resultado.grid(column=0, row=1)
Input = Entry(tab2, width=18)
Input.grid(column=1,row=0)
Output = Label(tab2,bg="white", fg="black",width= 18, font='Helvetica 9 bold')
Output.grid(column=1,row=1)
Valor = Label(tab2, text="€", bg="white", fg="black", width=10, font='Helvetica 9 bold')
Valor.grid(column=2,row=0)
Calcular = Button(tab2, text="Invertir", command=inversion,width=10, font='Helvetica 9 bold')
Calcular.grid(column=2, row=1)
Vender = Label(tab2, text="Vender", bg="pink", fg="black", width=18, font='Helvetica 9 bold')
Vender.grid(column=0, row=2)
Inputv = Entry(tab2, width=18)
Inputv.grid(column=1,row=2)
Resultado2 = Label(tab2, text="Resultado Venta" , bg="pink" , fg="black",width=18, font='Helvetica 9 bold')
Resultado2.grid(column=0, row=3)
Output2 = Label(tab2,bg="white", fg="black",width= 18, font='Helvetica 9 bold')
Output2.grid(column=1,row=3)
Simbolo_C = Label(tab2, text=variable.get(), bg="white", fg="black", width=10, font='Helvetica 9 bold')
Simbolo_C.grid(column=2, row=2)
CalcularV = Button(tab2, text="Vender", command=vender,width=10, font='Helvetica 9 bold')
CalcularV.grid(column=2, row=3)
Limpiar = Button(tab2, text="Limpiar", command=Limpiar, width=10, font='Helvetica 9 bold')
Limpiar.place(x=510,y=0)

Cartera = Label(tab2, text="Monedero" , bg="red" , fg="black",width= 18, font='Helvetica 9 bold')
Cartera.grid(column=0, row=4)
Balance = Label(tab2, text="Balance" , bg="sky blue" , fg="black",width= 18, font='Helvetica 9 bold')
Balance.grid(column=0, row=5)
BalanceOut = Label(tab2, bg="white", fg="black",width= 18, font='Helvetica 9 bold')
BalanceOut.grid(column=1, row=5)
BalanceOut.configure(text=monedero)
InputM = Entry(tab2, width=18)
InputM.grid(column=1,row=4)

AñadirSaldo = Button(tab2, text="Añadir Saldo", command=Añadir, width=10, font='Helvetica 9 bold')
AñadirSaldo.grid(column=2, row=4)

tab_control.pack(expand=1, fill='both')
window.mainloop()